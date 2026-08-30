"""
Acceso de MiClaw a tu PC: leer/escribir archivos y ejecutar comandos.

SEGURIDAD: ninguna herramienta se ejecuta sin tu confirmación. El asistente
propone la acción; el frontend te muestra una tarjeta con el comando/ruta y
tú la apruebas o la rechazas. Además, la escritura y la terminal solo operan
dentro de rutas permitidas (por defecto tu carpeta de usuario y la carpeta
del proyecto; ajustable en Ajustes → PC).

El token @@PC:...@@ es el "contrato" entre el agente y la interfaz:
  - @@PC:ver@@           → ver un archivo (se aprueba solo si es pequeño)
  - @@PC:escribir@@       → crear/sobrescribir un archivo (pedir permiso)
  - @@PC:terminal@@       → ejecutar un comando (pedir permiso)
  - @@PC:apuntes@@        → lista de tus apuntes (memoria) — sin permiso
"""

import os
import subprocess
import tempfile
from pathlib import Path

from . import config

# ---------------------------------------------------------------------------
# Rutas permitidas (sandbox)
# ---------------------------------------------------------------------------
def rutas_permitidas():
    """Devuelve la lista de directorios donde se permite escribir/ejecutar."""
    raiz = Path(config.BASE_DIR)
    custom = (config.load_config().get("pc") or {}).get("carpeta_extra", "")
    rutas = [raiz, Path.home()]
    if custom:
        p = Path(custom).expanduser()
        if p.is_dir():
            rutas.append(p)
    return rutas


def _esta_permitido(path):
    try:
        target = Path(path).expanduser().resolve()
    except OSError:
        return False
    for base in rutas_permitidas():
        try:
            target.relative_to(base.resolve())
            return True
        except ValueError:
            continue
    return False


def _info_error_permiso():
    return ("Acción bloqueada: la ruta está fuera de las carpetas permitidas "
            "(tu carpeta de usuario o la del proyecto).")


# ---------------------------------------------------------------------------
# Herramientas (llamadas DESPUÉS de la aprobación del usuario)
# ---------------------------------------------------------------------------
def ver_archivo(ruta, max_chars=200_000):
    try:
        path = Path(ruta).expanduser()
        if not path.is_file():
            return f"No encuentro el archivo: {ruta}"
        tamano = path.stat().st_size
        if tamano > max_chars:
            return f"El archivo es muy grande ({tamano} bytes). Ábrelo tú mismo."
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            contenido = f.read(max_chars)
        return (f"Contenido de {path}:\n\n{contenido}"[:15_000])
    except OSError as e:
        return f"No pude leer el archivo: {e}"


def listar_carpeta(ruta):
    """Lista el contenido de una carpeta (nombre, tipo, tamaño, fecha)."""
    try:
        path = Path(ruta).expanduser()
        if not path.is_dir():
            return f"No encuentro la carpeta: {ruta}"
        lineas = [f"Contenido de {path}:"]
        for item in sorted(path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower())):
            if item.is_dir():
                lineas.append(f"  📁 {item.name}/")
            else:
                tam = item.stat().st_size
                if tam >= 1024 * 1024:
                    tam_s = f"{tam / 1024 / 1024:.1f} MB"
                elif tam >= 1024:
                    tam_s = f"{tam / 1024:.0f} KB"
                else:
                    tam_s = f"{tam} B"
                lineas.append(f"  📄 {item.name}  ({tam_s})")
        return "\n".join(lineas[:300])
    except OSError as e:
        return f"No pude listar la carpeta: {e}"


def buscar_en(ruta, texto):
    """Busca texto dentro de archivos de una carpeta (grep simple)."""
    if not _esta_permitido(ruta):
        return _info_error_permiso()
    texto_low = texto.lower()
    resultados = []
    base = Path(ruta).expanduser()
    if not base.is_dir():
        return f"No encuentro la carpeta: {ruta}"
    TIPOS = {".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css", ".md", ".txt",
             ".json", ".csv", ".log", ".yml", ".yaml", ".toml", ".ini", ".sh",
             ".bat", ".c", ".cpp", ".h", ".java", ".go", ".rs", ".rb", ".php"}
    for path in base.rglob("*"):
        if path.is_dir() or path.name.startswith("."):
            continue
        if path.suffix.lower() not in TIPOS:
            continue
        if path.stat().st_size > 300_000:
            continue
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                for num, linea in enumerate(f, 1):
                    if texto_low in linea.lower():
                        resultados.append(f"{path}:{num}: {linea.rstrip()[:200]}")
                        if len(resultados) >= 60:
                            resultados.append("… (más resultados omitidos)")
                            return "\n".join(resultados)
        except OSError:
            continue
    if not resultados:
        return f"No encontré «{texto}» en {base}"
    return "\n".join(resultados)


def leer_documento(ruta):
    """Extrae texto de PDF, Word o Excel (si están instalados los paquetes)."""
    path = Path(ruta).expanduser()
    if not path.is_file():
        return f"No encuentro el archivo: {ruta}"
    suf = path.suffix.lower()
    try:
        if suf == ".pdf":
            from pypdf import PdfReader
            r = PdfReader(str(path))
            texto = "\n".join((p.extract_text() or "") for p in r.pages[:40])
            return f"Contenido de {path.name} ({len(r.pages)} págs):\n\n{texto[:15000]}"
        if suf == ".docx":
            import docx
            d = docx.Document(str(path))
            texto = "\n".join(p.text for p in d.paragraphs)
            return f"Contenido de {path.name}:\n\n{texto[:15000]}"
        if suf == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True)
            partes = []
            for hoja in wb.sheetnames[:3]:
                ws = wb[hoja]
                filas = []
                for fila in ws.iter_rows(values_only=True):
                    filas.append(" | ".join("" if c is None else str(c) for c in fila))
                    if len(filas) >= 100:
                        filas.append("…")
                        break
                partes.append(f"[Hoja: {hoja}]\n" + "\n".join(filas))
            return f"Contenido de {path.name}:\n\n" + "\n\n".join(partes)[:15000]
        return "Solo sé leer .pdf, .docx y .xlsx con esta herramienta."
    except ImportError:
        return ("Faltan paquetes para leer documentos. "
                "Ejecuta: pip install pypdf python-docx openpyxl")
    except Exception as e:
        return f"No pude leer el documento: {e}"


def escribir_archivo(ruta, contenido):
    if not _esta_permitido(ruta):
        return _info_error_permiso()
    try:
        path = Path(ruta).expanduser()
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(contenido)
        return f"Archivo guardado: {path}"
    except OSError as e:
        return f"No pude escribir el archivo: {e}"


def ejecutar_comando(comando):
    if not _esta_permitido("."):
        return _info_error_permiso()
    try:
        proc = subprocess.run(
            comando, shell=True, capture_output=True, text=True,
            timeout=120, cwd=config.BASE_DIR,
            env={**os.environ, "PYTHONIOENCODING": "utf-8"},
        )
        salida = (proc.stdout or "") + (proc.stderr or "")
        codigo = proc.returncode
        resumen = f"Salida (código {codigo}):\n{salida}" if salida else f"Comando terminado (código {codigo}) sin salida."
        return resumen[:10_000]
    except subprocess.TimeoutExpired:
        return "El comando tardó más de 120 segundos y se canceló."
    except OSError as e:
        return f"No pude ejecutar el comando: {e}"


# ---------------------------------------------------------------------------
# Parseo del token @@PC:...@@
# ---------------------------------------------------------------------------
def parsear(texto):
    """
    Devuelve (accion, datos) si el texto empieza por @@PC:accion@@.
    Para 'escribir' y 'terminal', los datos pueden ir en el propio texto
    (si el modelo usa líneas 'RUTA:'/'COMANDO:') o en la siguiente línea.
    """
    texto = texto.strip()
    if not texto.startswith("@@"):
        return None, None
    fin = texto.find("@@", 2)
    if fin == -1:
        return None, None
    token = texto[2:fin].strip()
    if not token.startswith("PC:"):
        return None, None
    accion = token[3:].strip()
    resto = texto[fin + 2:].strip()
    if accion not in ("ver", "escribir", "terminal", "apuntes", "listar", "buscar", "documento"):
        return None, None

    if accion == "ver":
        return accion, resto or None

    if accion == "listar":
        return accion, resto or None

    if accion == "documento":
        return accion, resto or None

    if accion == "buscar":
        ruta, texto = None, None
        for linea in resto.splitlines():
            l = linea.strip()
            if l.lower().startswith("ruta:"):
                ruta = l.split(":", 1)[1].strip()
            elif l.lower().startswith("texto:"):
                texto = l.split(":", 1)[1].strip()
        return accion, {"ruta": ruta, "texto": texto}

    if accion == "escribir":
        ruta = None
        contenido = resto
        for linea in resto.splitlines():
            if linea.lower().startswith("ruta:"):
                ruta = linea.split(":", 1)[1].strip()
                contenido = "\n".join(resto.splitlines()[1:])
                break
        if ruta is None:
            # Si no hay 'RUTA:', el primer bloque no vacío es la ruta y el
            # resto (tras un salto de línea) es el contenido.
            partes = resto.split("\n", 1)
            if len(partes) == 2 and partes[0].strip():
                ruta, contenido = partes[0].strip(), partes[1].strip()
        return accion, {"ruta": ruta, "contenido": contenido.strip()}

    if accion == "terminal":
        comando = resto
        for linea in resto.splitlines():
            if linea.lower().startswith("comando:"):
                comando = linea.split(":", 1)[1].strip()
                break
        return accion, comando

    return accion, None


def ejecutar(accion, datos):
    """Ejecuta la acción YA aprobada por el usuario."""
    if accion == "ver":
        return ver_archivo(datos or "")
    if accion == "escribir":
        if not datos or not datos.get("ruta"):
            return "Falta la ruta del archivo."
        return escribir_archivo(datos["ruta"], datos.get("contenido", ""))
    if accion == "terminal":
        return ejecutar_comando(datos or "")
    if accion == "apuntes":
        from . import memory
        return memory.read_memory() or "(sin apuntes)"
    if accion == "listar":
        return listar_carpeta(datos or "")
    if accion == "buscar":
        if not datos or not datos.get("ruta") or not datos.get("texto"):
            return "Faltan RUTA o TEXTO para buscar."
        return buscar_en(datos["ruta"], datos["texto"])
    if accion == "documento":
        return leer_documento(datos or "")
    return "Acción desconocida."


def parsear_lote(texto):
    """
    Extrae TODAS las acciones @@PC:...@@ de un texto (para aprobarlas en lote).
    Devuelve una lista de (accion, datos). Vacía si no hay ninguna.
    """
    import re
    piezas = re.split(r"(?m)(?=^@@PC:[a-z]+@@)", texto.strip())
    acciones = []
    for pieza in piezas:
        accion, datos = parsear(pieza)
        if accion:
            acciones.append({"accion": accion, "datos": datos})
    return acciones
